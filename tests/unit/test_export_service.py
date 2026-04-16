"""Unit tests for ExportService (CSV and XLSX export)."""
import csv
import io
from datetime import datetime, timedelta

import pytest
import pytest_asyncio

from bot.models.models import Category, Status
from bot.repositories.request_repo import create_request, update_status
from bot.repositories.user_repo import get_or_create_user
from bot.services.export_service import COLUMNS, ExportService


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

async def _make_user(session, telegram_id: int = 1):
    return await get_or_create_user(session, telegram_id=telegram_id, username="tester")


async def _make_request(session, user_id: int, **kwargs):
    defaults = dict(
        category=Category.LOST,
        description="A lost dog near the park",
        location={"latitude": 50.4, "longitude": 30.5, "address_text": "Kyiv"},
        contact="@tester",
    )
    defaults.update(kwargs)
    return await create_request(session, user_id=user_id, **defaults)


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_export_csv_returns_bytes(session):
    svc = ExportService(session)
    result = await svc.export_csv()
    assert isinstance(result, bytes)


@pytest.mark.asyncio
async def test_export_csv_header_row(session):
    svc = ExportService(session)
    result = await svc.export_csv()
    reader = csv.reader(io.StringIO(result.decode("utf-8")))
    header = next(reader)
    assert header == COLUMNS


@pytest.mark.asyncio
async def test_export_csv_empty_no_data_rows(session):
    svc = ExportService(session)
    result = await svc.export_csv()
    reader = csv.reader(io.StringIO(result.decode("utf-8")))
    rows = list(reader)
    # Only the header row
    assert len(rows) == 1


@pytest.mark.asyncio
async def test_export_csv_one_request(session):
    user = await _make_user(session)
    req = await _make_request(session, user_id=user.id)

    svc = ExportService(session)
    result = await svc.export_csv()
    reader = csv.reader(io.StringIO(result.decode("utf-8")))
    rows = list(reader)

    assert len(rows) == 2  # header + 1 data row
    data = rows[1]
    assert data[0] == str(req.id)
    assert data[1] == Category.LOST.value
    assert data[3] == Status.NEW.value
    assert data[6] == str(30.5)
    assert data[7] == "Kyiv"
    assert data[8] == "@tester"


@pytest.mark.asyncio
async def test_export_csv_multiple_requests(session):
    user = await _make_user(session)
    await _make_request(session, user_id=user.id, category=Category.LOST)
    await _make_request(session, user_id=user.id, category=Category.INJURED)

    svc = ExportService(session)
    result = await svc.export_csv()
    reader = csv.reader(io.StringIO(result.decode("utf-8")))
    rows = list(reader)
    assert len(rows) == 3  # header + 2 data rows


@pytest.mark.asyncio
async def test_export_csv_filter_by_category(session):
    user = await _make_user(session)
    await _make_request(session, user_id=user.id, category=Category.LOST)
    await _make_request(session, user_id=user.id, category=Category.INJURED)

    svc = ExportService(session)
    result = await svc.export_csv(filters={"category": "LOST"})
    reader = csv.reader(io.StringIO(result.decode("utf-8")))
    rows = list(reader)
    assert len(rows) == 2
    assert rows[1][1] == "LOST"


@pytest.mark.asyncio
async def test_export_csv_filter_by_status(session):
    user = await _make_user(session)
    req = await _make_request(session, user_id=user.id)
    await update_status(session, req.id, Status.IN_PROGRESS)
    await _make_request(session, user_id=user.id)  # stays NEW

    svc = ExportService(session)
    result = await svc.export_csv(filters={"status": "IN_PROGRESS"})
    reader = csv.reader(io.StringIO(result.decode("utf-8")))
    rows = list(reader)
    assert len(rows) == 2
    assert rows[1][3] == "IN_PROGRESS"


@pytest.mark.asyncio
async def test_export_csv_filter_by_date_range(session):
    user = await _make_user(session)
    req = await _make_request(session, user_id=user.id)

    now = datetime.utcnow()
    svc = ExportService(session)
    result = await svc.export_csv(filters={
        "date_from": now - timedelta(minutes=1),
        "date_to": now + timedelta(minutes=1),
    })
    reader = csv.reader(io.StringIO(result.decode("utf-8")))
    rows = list(reader)
    assert any(row[0] == str(req.id) for row in rows[1:])


@pytest.mark.asyncio
async def test_export_csv_no_location(session):
    user = await _make_user(session)
    await create_request(
        session,
        user_id=user.id,
        category=Category.DEAD,
        description="Dead animal on road",
        location=None,
        contact=None,
    )

    svc = ExportService(session)
    result = await svc.export_csv()
    reader = csv.reader(io.StringIO(result.decode("utf-8")))
    rows = list(reader)
    data = rows[1]
    # latitude and longitude columns should be empty strings
    assert data[5] == "None" or data[5] == ""
    assert data[6] == "None" or data[6] == ""


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_export_xlsx_returns_bytes(session):
    svc = ExportService(session)
    result = await svc.export_xlsx()
    assert isinstance(result, bytes)


@pytest.mark.asyncio
async def test_export_xlsx_valid_workbook(session):
    import openpyxl

    svc = ExportService(session)
    result = await svc.export_xlsx()

    wb = openpyxl.load_workbook(io.BytesIO(result))
    assert "Requests" in wb.sheetnames


@pytest.mark.asyncio
async def test_export_xlsx_header_row(session):
    import openpyxl

    svc = ExportService(session)
    result = await svc.export_xlsx()

    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["Requests"]
    header = [cell.value for cell in ws[1]]
    assert header == COLUMNS


@pytest.mark.asyncio
async def test_export_xlsx_empty_only_header(session):
    import openpyxl

    svc = ExportService(session)
    result = await svc.export_xlsx()

    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["Requests"]
    assert ws.max_row == 1


@pytest.mark.asyncio
async def test_export_xlsx_one_request(session):
    import openpyxl

    user = await _make_user(session)
    req = await _make_request(session, user_id=user.id)

    svc = ExportService(session)
    result = await svc.export_xlsx()

    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["Requests"]
    assert ws.max_row == 2
    data_row = [cell.value for cell in ws[2]]
    assert data_row[0] == req.id
    assert data_row[1] == Category.LOST.value
    assert data_row[3] == Status.NEW.value


@pytest.mark.asyncio
async def test_export_xlsx_filter_by_category(session):
    import openpyxl

    user = await _make_user(session)
    await _make_request(session, user_id=user.id, category=Category.LOST)
    await _make_request(session, user_id=user.id, category=Category.INJURED)

    svc = ExportService(session)
    result = await svc.export_xlsx(filters={"category": Category.INJURED})

    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["Requests"]
    assert ws.max_row == 2  # header + 1 row
    assert ws.cell(2, 2).value == "INJURED"


@pytest.mark.asyncio
async def test_export_xlsx_none_filters(session):
    """Passing None filters should export all requests."""
    import openpyxl

    user = await _make_user(session)
    await _make_request(session, user_id=user.id)
    await _make_request(session, user_id=user.id)

    svc = ExportService(session)
    result = await svc.export_xlsx(filters=None)

    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["Requests"]
    assert ws.max_row == 3  # header + 2 rows
